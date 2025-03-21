import os
from datetime import datetime
import re
from openpyxl import load_workbook, Workbook

# Paths
log_folder = "logs"  # Folder where .txt files are located

current_time = datetime.now().strftime("%Y%m%d_%H%M%S")  # Format: YYYYMMDD_HHMMSS
output_file = f"Report_{current_time}.xlsx"

# Column names in the Excel file (must match exactly with your sheet, updated to lowercase)
columns = [
    "SrNo", "DeviceSerialNumber", "RRNumber", "TrxStart", "CardDecryptionReq", "CardDecryptionRes",
    "MackingReq", "MackingRes", "RequestToSP", "ResponseFromSP", "TrxEnd", "TotalTime", "TotalTimeW/OSP"
]

# Columns to leave blank (empty strings)
blank_columns = ["RRNumber"]

# Matching string
matchingTerminalString = "20049729"

# Field mapping for regex extraction (field names to match in log files)
patterns = {
    "TrxStart": re.compile(r'Request Received: Sale'),
    "DeviceSerialNumber": re.compile(r'"deviceSerialNo":"([^"]+)"'),  # Extract from "Transaction Started" line
    "RRNumber": re.compile(r'"rrNumber":"([^"]+)"'),
    "CardDecryptionReq": re.compile(r'Request Sent to HSM for Card details'),
    "CardDecryptionRes": re.compile(r'Decryption of Card details successfully'),
    "MackingReq": re.compile(r'Request Sent to HSM for Macking'),
    "MackingRes": re.compile(r'Macking successfully'),
    "RequestToSP": re.compile(r'ISO Parsed message Send Request Length'),
    "ResponseFromSP": re.compile(r'ConnectionFileAppender - ISO Parsed message Received Response Length'),
    "TrxEnd": re.compile(r'Transaction End \((\d+)\)'),
}

timePattern = re.compile(r"(\d{2}:\d{2}:\d{2}),(\d{3})")

# Helper function to calculate TotalTime and TotalTimeW/OSP
def calculate_times(row):
    # Convert timestamps to datetime objects
    def parse_time(time_str):
        if not time_str:
            return None
        return datetime.strptime(time_str, "%H:%M:%S.%f")

    # Extract timestamps from the row
    trx_start = parse_time(row.get("TrxStart"))
    request_to_sp = parse_time(row.get("RequestToSP"))
    response_from_sp = parse_time(row.get("ResponseFromSP"))
    trx_end = parse_time(row.get("TrxEnd"))

    # Calculate TotalTimeW/OSP
    if trx_start and request_to_sp and response_from_sp and trx_end:
        total_time_w_osp = (trx_end - trx_start) - (response_from_sp - request_to_sp)
    else:
        total_time_w_osp = None

    # Calculate TotalTime
    if trx_start and trx_end:
        total_time = trx_end - trx_start
    else:
        total_time = None

    # Format results as strings
    return (
        str(total_time) if total_time else "",
        str(total_time_w_osp) if total_time_w_osp else ""
    )

# Helper function to process the log file
srNo = 1  # Initialize the serial number counter
def process_log_file(file_path):
    global srNo  # Declare srNo as global to modify its value across function calls

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:  # Use 'ignore' to skip errors
        lines = file.readlines()

    data = []
    processed_serial_numbers = set()  # Keep track of processed DeviceSerialNumbers
    row = {}  # A single row to accumulate data between markers
    writing_started = False  # Flag to indicate when data should be written

    # Updated time pattern to include milliseconds
    timePattern = re.compile(r"(\d{2}:\d{2}:\d{2}),(\d{3})")

    for line in lines:
        # Check for "Transaction Started" marker
        start_match = re.search(patterns["TrxStart"], line)
        if start_match:
            writing_started = True  # Start processing data
            row = {"SrNo": srNo}  # Initialize a new row with the serial number
            srNo += 1  # Increment serial number for the next row
            
            # Extract timestamp for "Transaction Started" if available
            time_match = timePattern.search(line)
            if time_match:
                hours_minutes_seconds = time_match.group(1)  # "HH:mm:ss"
                milliseconds = time_match.group(2)  # "563"
                formatted_time = f"{hours_minutes_seconds}.{milliseconds}"  # Replace ',' with '.'
                row["TrxStart"] = formatted_time  # Add the formatted timestamp
            
            continue

        device_number = re.search(patterns["DeviceSerialNumber"], line)
        if device_number:
            device_serial_no = device_number.group(1)
            row["DeviceSerialNumber"] = device_serial_no  # Add DeviceSerialNumber to the row
            continue

        # Process data if writing has started
        if writing_started:
            for column, pattern in patterns.items():
                if column in ["DeviceSerialNumber", "TrxStart"]:
                    continue  # Skip already processed fields

                match = re.search(pattern, line)
                if match:
                    # Extract and store the matched value in the row
                    if column in row:
                        continue  # Avoid overwriting existing data

                    if column in ["RRNumber"]:
                        row[column] = match.group(1)  # Add the matched value directly
                    else:
                        time_match = timePattern.search(line)  # Extract time
                        if time_match:
                            hours_minutes_seconds = time_match.group(1)  # "HH:mm:ss"
                            milliseconds = time_match.group(2)  # "563"
                            formatted_time = f"{hours_minutes_seconds}.{milliseconds}"  # Replace ',' with '.'
                            row[column] = formatted_time  # Add the formatted time to the row

        # Check for "Transaction End" marker
        if "Transaction End" in line:
            writing_started = False  # Stop processing data

            # Skip rows without DeviceSerialNumber or duplicate rows
            device_serial_no = row.get("DeviceSerialNumber", "")
            if not device_serial_no or device_serial_no in processed_serial_numbers:
                row = {}  # Reset the row
                continue

            # Calculate TotalTime and TotalTimeW/OSP for the row
            total_time, total_time_w_osp = calculate_times(row)
            row["TotalTime"] = total_time
            row["TotalTimeW/OSP"] = total_time_w_osp

            # Add the completed row to the data list
            data.append(row)
            processed_serial_numbers.add(device_serial_no)  # Mark serial number as processed
            row = {}  # Reset the row for the next transaction

    return data



# Function to write data to Excel
def write_to_excel(data, output_file):
    # Load the existing workbook or create a new one if it doesn't exist
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Transaction Data"
    
    # Write the header row (columns)
    for col_idx, col in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col)  # Writing the header in row 1
    
    # Write the data starting from row 2 (skip the header row)
    start_row = 2  # We want to start inserting data from row 2
    for idx, row in enumerate(data, start=start_row):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=idx, column=col_idx, value=row.get(col, ""))
    
    # Save the workbook (it won't overwrite the headers)
    wb.save(output_file)
    print(f"Data written to {output_file}")


# Main code to process logs and create Excel report
def main():
    all_data = []
    for file_name in os.listdir(log_folder):
        if file_name.endswith(".txt"):
            file_path = os.path.join(log_folder, file_name)
            file_data = process_log_file(file_path)
            all_data.extend(file_data)

    write_to_excel(all_data, output_file)


if __name__ == "__main__":
    main()
