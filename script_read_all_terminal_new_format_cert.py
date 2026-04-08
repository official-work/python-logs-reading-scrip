import os
from datetime import datetime
import re
from openpyxl import load_workbook, Workbook

#Create folder if not exists
os.makedirs("TerminalExports", exist_ok=True)

# Paths
log_folder = "logs"  # Folder where .txt files are located

current_time = str(int(datetime.now().timestamp() * 1000))
#output_file = f"Report_{current_time}.xlsx"
output_file = os.path.join("TerminalExports", f"All_Terminal_Report_{current_time}.xlsx")

# Column names in the Excel file (must match exactly with your sheet, updated to lowercase)
columns = [
    "SrNo", "TxnType", "SubTxnType", "DeviceSerialNumber", "RRNumber", "TrxStart", "CardDecryptionReq", "CardDecryptionRes",
    "MackingReq", "MackingRes", "RequestToSP", "ResponseFromSP", "TrxEnd", "TotalTime", "TotalTimeW/OSP"
]

# Field mapping for regex extraction (field names to match in log files)
patterns = {
    #"TrxStart": re.compile(r'Request Received: Sale'),
    "TrxStart": re.compile(r'Request Received: '), # Adjusted to match any transaction type with Prasana
    "TxnType": re.compile(r'"txnType":"([^"]+)"'), # TxnType can be Sale, Refund, Void, etc.
    "SubTxnType": re.compile(r'"subTxnType":"([^"]+)"'), # This will match any transaction type
    "DeviceSerialNumber": re.compile(r'"deviceSerialNo":"([^"]+)"'),  # Extract from "Transaction Started" line
    "RRNumber": re.compile(r'"rrNumber":"([^"]+)"'),
    "CardDecryptionReq": re.compile(r'Request Sent to HSM for Card details'),
    "CardDecryptionRes": re.compile(r'Decryption of Card details successfully'),
    "MackingReq": re.compile(r'Request Sent to HSM for Macking'),
    "MackingRes": re.compile(r'Macking successfully'),
    "RequestToSP": re.compile(r'ISO Parsed message Send Request Length'),
    "ResponseFromSP": re.compile(r'ConnectionFileAppender - ISO Parsed message Received Response Length'),
    "TrxEnd": re.compile(r'Transaction End \((\d+)\)'),
}

timePattern = re.compile(r"(\d{2}:\d{2}:\d{2}),(\d{3})")
UUID_PATTERN = re.compile(r'\[([a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12})\]')

# Helper function to calculate TotalTime and TotalTimeW/OSP
def calculate_times(row):
    # Convert timestamps to datetime objects
    def parse_time(time_str):
        if not time_str:
            return None
        return datetime.strptime(time_str, "%H:%M:%S.%f")

    # Extract timestamps from the row
    trx_start = parse_time(row.get("TrxStart"))
    request_to_sp = parse_time(row.get("RequestToSP"))
    response_from_sp = parse_time(row.get("ResponseFromSP"))
    trx_end = parse_time(row.get("TrxEnd"))

    # Calculate TotalTimeW/OSP
    if trx_start and request_to_sp and response_from_sp and trx_end:
        total_time_w_osp = (trx_end - trx_start) - (response_from_sp - request_to_sp)
    else:
        total_time_w_osp = None

    # Calculate TotalTime
    if trx_start and trx_end:
        total_time = trx_end - trx_start
    else:
        total_time = None

    # Format results as strings
    return (
        str(total_time) if total_time else "",
        str(total_time_w_osp) if total_time_w_osp else ""
    )


# Process all log lines belonging to a single transaction UUID
def process_transaction(lines):
    row = {}
    has_trx_start = False

    for line in lines:
        # TxnType / SubTxnType — first match wins
        if "TxnType" not in row:
            m = re.search(patterns["TxnType"], line)
            if m:
                row["TxnType"] = m.group(1)

        if "SubTxnType" not in row:
            m = re.search(patterns["SubTxnType"], line)
            if m:
                row["SubTxnType"] = m.group(1)

        # DeviceSerialNumber — first match wins
        if "DeviceSerialNumber" not in row:
            m = re.search(patterns["DeviceSerialNumber"], line)
            if m:
                row["DeviceSerialNumber"] = m.group(1)

        # RRNumber — first match wins (covers Request JSON, Response JSON, and post-TrxEnd lines)
        if "RRNumber" not in row:
            m = re.search(patterns["RRNumber"], line)
            if m:
                row["RRNumber"] = m.group(1)

        # TrxStart
        if re.search(patterns["TrxStart"], line):
            has_trx_start = True
            t = timePattern.search(line)
            if t:
                row["TrxStart"] = f"{t.group(1)}.{t.group(2)}"

        # Time-stamped fields — first match wins
        for col in ["CardDecryptionReq", "CardDecryptionRes", "MackingReq",
                    "MackingRes", "RequestToSP", "ResponseFromSP", "TrxEnd"]:
            if col not in row:
                m = re.search(patterns[col], line)
                if m:
                    t = timePattern.search(line)
                    if t:
                        row[col] = f"{t.group(1)}.{t.group(2)}"

    if not has_trx_start or not row.get("DeviceSerialNumber"):
        return None

    total_time, total_time_w_osp = calculate_times(row)
    row["TotalTime"] = total_time
    row["TotalTimeW/OSP"] = total_time_w_osp
    return row


# Helper function to process the log file
srNo = 1  # Initialize the serial number counter
def process_log_file(file_path):
    global srNo  # Declare srNo as global to modify its value across function calls

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
        lines = file.readlines()

    # Group lines by transaction UUID to handle multi-threaded interleaving
    transactions = {}       # uuid -> list of lines (preserves line order)
    first_timestamps = {}   # uuid -> first seen timestamp (for chronological sort)

    for line in lines:
        m = UUID_PATTERN.search(line)
        if not m:
            continue
        uuid = m.group(1)
        if uuid not in transactions:
            transactions[uuid] = []
            t = timePattern.search(line)
            if t:
                first_timestamps[uuid] = f"{t.group(1)}.{t.group(2)}"
        transactions[uuid].append(line)

    # Sort UUIDs by first-seen timestamp to preserve chronological order
    sorted_uuids = sorted(transactions, key=lambda u: first_timestamps.get(u, ""))

    data = []
    for uuid in sorted_uuids:
        row = process_transaction(transactions[uuid])
        if row:
            row["SrNo"] = srNo
            srNo += 1
            data.append(row)

    return data



# Function to write data to Excel
def write_to_excel(data, output_file):
    # Load the existing workbook or create a new one if it doesn't exist
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Transaction Data"

    # Write the header row (columns)
    for col_idx, col in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col)  # Writing the header in row 1

    # Write the data starting from row 2 (skip the header row)
    start_row = 2  # We want to start inserting data from row 2
    for idx, row in enumerate(data, start=start_row):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=idx, column=col_idx, value=row.get(col, ""))

    # Save the workbook (it won't overwrite the headers)
    wb.save(output_file)
    print(f"Data written to {output_file}")


# Main code to process logs and create Excel report
def main():
    all_data = []
    for file_name in os.listdir(log_folder):
        if file_name.endswith(".txt"):
            file_path = os.path.join(log_folder, file_name)
            file_data = process_log_file(file_path)
            all_data.extend(file_data)

    write_to_excel(all_data, output_file)


if __name__ == "__main__":
    main()
