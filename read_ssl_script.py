import os
import re
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

# Create folder if not exists
os.makedirs("Exports", exist_ok=True)

# Configuration
log_folder = "logs"
current_time = str(int(datetime.now().timestamp() * 1000))
output_file = os.path.join("Exports", f"SSLFingerprint_Report_{current_time}.xlsx")

# Columns in Excel
columns = ["SrNo", "FilePath", "ResponseTime", "UserName", "LastModifiedDate", "ResponseData"]

# Regex patterns
request_pattern = re.compile(r"GetSSLFingerprint Request received:([^:]+):({.*})")
response_pattern = re.compile(r"GetSSLFingerprint Response sent:([^:]+):({.*})")

# Updated regex to extract full datetime with milliseconds
dateTimePattern = re.compile(r"(\d{2} \w{3} \d{4} \d{2}:\d{2}:\d{2}),(\d{3})")

# Process log files
def process_requests(file_path, user_data, sr_counter):
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        for line in f:
            # Handle request line
            request_match = request_pattern.search(line)
            if request_match:
                username, json_str = request_match.groups()
                try:
                    json_data = json.loads(json_str)
                    last_modified = json_data.get("lastModifiedDate", "")
                except json.JSONDecodeError:
                    last_modified = ""

                if username in user_data:
                    user_data[username]["LastModifiedDate"] = last_modified
                else:
                    user_data[username] = {
                        "SrNo": sr_counter,
                        "FilePath": file_path,
                        "UserName": username,
                        "LastModifiedDate": last_modified,
                        "ResponseTime": "",
                        "ResponseData": ""
                    }
                    sr_counter += 1

            # Handle response line
            response_match = response_pattern.search(line)
            if response_match:
                username, json_str = response_match.groups()

                # Extract response timestamp
                response_datetime = ''
                time_match = dateTimePattern.search(line)
                if time_match:
                    raw_datetime = f"{time_match.group(1)},{time_match.group(2)}"  # "08 May 2025 13:23:46,024"
                    try:
                        dt_obj = datetime.strptime(raw_datetime, "%d %b %Y %H:%M:%S,%f")
                        response_datetime = dt_obj.strftime("%Y-%m-%d %H:%M:%S")  # Standard format
                    except ValueError:
                        response_datetime = raw_datetime  # Fallback in case of parsing issue

                if username in user_data:
                    user_data[username]["ResponseTime"] = response_datetime
                    user_data[username]["ResponseData"] = json_str

    return sr_counter

# Write to Excel
def write_to_excel(user_data, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fingerprint Requests"

    # Header
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = Font(bold=True)

    # Rows
    for idx, user in enumerate(user_data.values(), start=2):
        for col_idx, col in enumerate(columns, 1):
            ws.cell(row=idx, column=col_idx, value=user.get(col, ""))

    wb.save(output_file)
    print(f"Report saved to {output_file}")

# Main function
def main():
    user_data = {}
    sr_counter = 1

    for file_name in os.listdir(log_folder):
        if file_name.endswith(".txt"):
            file_path = os.path.join(log_folder, file_name)
            sr_counter = process_requests(file_path, user_data, sr_counter)

    write_to_excel(user_data, output_file)

if __name__ == "__main__":
    main()
