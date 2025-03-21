import os
from datetime import datetime
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

# Paths
log_folder = "logs"  # Folder where .txt files are located

current_time = datetime.now().strftime("%Y%m%d_%H%M%S")  # Format: YYYYMMDD_HHMMSS
output_file = f"Report_{current_time}.xlsx"

# Column names in the Excel file (must match exactly with your sheet, updated to lowercase)
columns = [
    "SrNo", "FilePath", "DeviceSerialNumber", "RRNumber", "TxnType", "TrxStart", "CardDecryptionReq", "CardDecryptionRes",
    "MackingReq", "MackingRes", "RequestToSP", "ResponseFromSP", "TrxEnd", "TotalTime", "TotalTimeW/OSP"
]


#Give terminal number for which get logs
matchingTerminalString = "20049907"

# Field mapping for regex extraction (field names to match in log files)
patterns = {
    "TrxStart": re.compile(rf'TransactionController - \({matchingTerminalString}\) Request Received: Sale'),
    "DeviceSerialNumber": re.compile(r'"deviceSerialNo":"([^"]+)"'),  # Extract from "Transaction Started" line
    "RRNumber": re.compile(r'"rrNumber":"([^"]+)"'),
    "TxnType": re.compile(r'"txnType":"([^"]+)"'),
    "CardDecryptionReq": re.compile(r'Request Sent to HSM for Card details'),
    "CardDecryptionRes": re.compile(r'Decryption of Card details successfully'),
    "MackingReq": re.compile(r'Request Sent to HSM for Macking'),
    "MackingRes": re.compile(r'Macking successfully'),
    "RequestToSP": re.compile(r'ISO Parsed message Send Request Length'),
    "ResponseFromSP": re.compile(r'ConnectionFileAppender - ISO Parsed message Received Response Length'),
    "TrxEnd": re.compile(r'Transaction End \((\d+)\)'),
}

timePattern = re.compile(r"(\d{2}:\d{2}:\d{2}),(\d{3})")


# Helper function to calculate TotalTime and TotalTimeW/OSP
def calculate_times(row):
    # Convert timestamps to datetime objects
    def parse_time(time_str):
        if not time_str:
            return None
        return datetime.strptime(time_str, "%H:%M:%S.%f")

    # Extract timestamps from the row
    trx_start = parse_time(row.get("TrxStart"))
    request_to_sp = parse_time(row.get("RequestToSP"))
    response_from_sp = parse_time(row.get("ResponseFromSP"))
    trx_end = parse_time(row.get("TrxEnd"))

    # Calculate TotalTimeW/OSP
    if trx_start and request_to_sp and response_from_sp and trx_end:
        total_time_w_osp = (trx_end - trx_start) - (response_from_sp - request_to_sp)
    else:
        total_time_w_osp = None

    # Calculate TotalTime
    if trx_start and trx_end:
        total_time = trx_end - trx_start
    else:
        total_time = None

    # Format results as strings
    return (
        str(total_time) if total_time else "",
        str(total_time_w_osp) if total_time_w_osp else ""
    )

# Helper function to process the log file
srNo = 1  # Initialize the serial number counter
def process_log_file(file_path):
    global srNo  # Declare srNo as global to modify its value across function calls

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:  # Use 'ignore' to skip errors
        lines = file.readlines()

    data = []
    processed_rr_numbers = set()  # Keep track of processed RRNumbers
    writing_started = False  # Flag to indicate when data should be written
    row = {}  # A single row to accumulate data between markers

    # Updated time pattern to include milliseconds
    timePattern = re.compile(r"(\d{2}:\d{2}:\d{2}),(\d{3})")

    for line in lines:
        # Check for "Transaction Started" marker specific to the terminal
        start_match = re.search(patterns["TrxStart"], line)
        if start_match:
            terminal_id = matchingTerminalString
            writing_started = True  # Start processing data
            row = {"SrNo": srNo}  # Initialize a new row with the serial number
            row['FilePath'] = file_path  # Initialize a new row with the file path
            srNo += 1  # Increment serial number for the next row
            
            # Extract timestamp for "Transaction Started" if available
            time_match = timePattern.search(line)
            if time_match:
                hours_minutes_seconds = time_match.group(1)  # "HH:mm:ss"
                milliseconds = time_match.group(2)  # "563"
                formatted_time = f"{hours_minutes_seconds}.{milliseconds}"  # Replace ',' with '.'
                row["TrxStart"] = formatted_time  # Add the formatted timestamp
            
            row["DeviceSerialNumber"] = terminal_id  # Add terminal ID to row
            continue

        # Extract RRNumber
        rr_number_match = re.search(patterns["RRNumber"], line)
        if rr_number_match:
            rr_number = rr_number_match.group(1)
            if rr_number in processed_rr_numbers:  # Skip duplicate RRNumbers
                writing_started = False
                row = {}
                continue
            row["RRNumber"] = rr_number  # Add RRNumber to the row
            processed_rr_numbers.add(rr_number)  # Mark as processed
            continue

        # Extract TxnType
        txn_type_match = re.search(patterns["TxnType"], line)
        if txn_type_match:
            txn_type = txn_type_match.group(1)
            row["TxnType"] = txn_type  # Add TxnType to the row
            continue

        # Process data if writing has started
        if writing_started:
            for column, pattern in patterns.items():
                if column in ["DeviceSerialNumber", "TrxStart", "RRNumber", "TxnType"]:
                    continue  # Skip already processed fields

                match = re.search(pattern, line)
                if match:
                    # Extract and store the matched value in the row
                    if column in row:
                        continue  # Avoid overwriting existing data

                    time_match = timePattern.search(line)  # Extract time
                    if time_match:
                        hours_minutes_seconds = time_match.group(1)  # "HH:mm:ss"
                        milliseconds = time_match.group(2)  # "563"
                        formatted_time = f"{hours_minutes_seconds}.{milliseconds}"  # Replace ',' with '.'
                        row[column] = formatted_time  # Add the formatted time to the row

        # Check for "Transaction End" marker specific to the terminal
        if f"Transaction End ({matchingTerminalString})" in line:
            writing_started = False  # Stop processing data

            # Skip rows without RRNumber or incomplete transactions
            if not row.get("RRNumber"):
                row = {}  # Reset the row
                continue

            # Calculate TotalTime and TotalTimeW/OSP for the row
            total_time, total_time_w_osp = calculate_times(row)
            row["TotalTime"] = total_time
            row["TotalTimeW/OSP"] = total_time_w_osp

            # Add the completed row to the data list
            data.append(row)
            row = {}  # Reset the row for the next transaction

    return data



# Function to write data to Excel
def write_to_excel(data, output_file):
    # Load the existing workbook or create a new one if it doesn't exist
    if os.path.exists(output_file):
        wb = load_workbook(output_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Transaction Data"
    
    # Write the header row (columns) with bold formatting
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)  # Writing the header in row 1
        cell.font = Font(bold=True)  # Set the font to bold
    
    # Write the data starting from row 2 (skip the header row)
    start_row = 2  # We want to start inserting data from row 2
    for idx, row in enumerate(data, start=start_row):
        for col_idx, col in enumerate(columns, start=1):
            ws.cell(row=idx, column=col_idx, value=row.get(col, ""))
    
    # Save the workbook (it won't overwrite the headers)
    wb.save(output_file)
    print(f"Data written to {output_file}")


# Main code to process logs and create Excel report
def main():
    all_data = []
    for file_name in os.listdir(log_folder):
        if file_name.endswith(".txt"):
            file_path = os.path.join(log_folder, file_name)
            #print(f"Log files: {file_path}")
            file_data = process_log_file(file_path)
            all_data.extend(file_data)

    write_to_excel(all_data, output_file)


if __name__ == "__main__":
    main()
